from pathlib import Path
import os
import pytz
from datetime import datetime
from django.test import TransactionTestCase
from django.core.management import call_command
from django.db.models import Count
from itertools import islice
from openpyxl import load_workbook

from comptages.test import utils as test_utils
from comptages.test import yearly_count_for
from comptages.datamodel import models
from comptages.report.yearly_report_bike import YearlyReportBike
from comptages.core import report, importer


class ReportTest(TransactionTestCase):
    @classmethod
    def setUpClass(cls):
        cls.test_outputs = Path("/OpenComptage/test_outputs")
        cls.test_data = Path("/test_data")

        for file in cls.test_outputs.iterdir():
            os.remove(file)

    def setUp(self):
        for file in Path(self.test_outputs).iterdir():
            os.remove(file)
        # With TransactionTestCase the db is reset at every test, so we
        # re-import base data every time.
        call_command("importdata")
        for file in Path(self.test_outputs).iterdir():
            os.remove(file)

    def test_report(self):
        # Create count and import some data
        installation = models.Installation.objects.get(name="00056520")
        count = yearly_count_for(2021, installation)
        importer.import_file(test_utils.test_data_path("00056520.V01"), count)
        importer.import_file(test_utils.test_data_path("00056520.V02"), count)
        report.prepare_reports("/tmp/", count)

    def test_generate_yearly_reports(self):
        call_command("importdata", "--only-count", "--limit", 100)
        # year and section id come from --add-count`
        year = 2021
        section_id = "00107695"
        installations = models.Installation.objects.filter(lane__id_section=section_id)
        installation = installations.first()
        assert installation

        count = models.Count.objects.filter(id_installation=installation.id).first()
        report.prepare_reports(
            self.test_outputs, count, year, "yearly", sections_ids=[section_id]
        )

        self.assertTrue(list(Path(self.test_outputs).iterdir()))

    def test_generate_yearly_reports_special_case(self):
        call_command("importdata", "--only-count", "--limit", 100)

        special_case_installations = models.Installation.objects.annotate(
            sections=Count("lane__id_section")
        ).filter(sections__lte=3)
        self.assertTrue(special_case_installations.exists())

        count = models.Count.objects.filter(
            id_installation__in=special_case_installations
        ).first()
        assert count

        year = count.start_process_date.year
        section = (
            models.Section.objects.filter(lane__id_installation=count.id_installation)
            .distinct()
            .first()
        )
        assert section

        print(
            f"Preparing reports for count = {count}, year = {year}, section_id = {section.id}"
        )
        report.prepare_reports(
            self.test_outputs, count, year, "yearly", sections_ids=[section.id]
        )
        self.assertTrue(list(Path(self.test_outputs).iterdir()))

    def test_all_sections_default(self):
        # Test if default report features all sections for special case
        test_data_folder = "5350_1_4"
        section_id = "53526896"

        installation = models.Installation.objects.get(lane__id_section_id=section_id)
        n_sections = (
            models.Lane.objects.filter(id_installation=installation.id)
            .values("id_section")
            .count()
        )
        self.assertGreater(n_sections, 0)

        model = models.Model.objects.all().first()
        device = models.Device.objects.all().first()
        sensor_type = models.SensorType.objects.all().first()
        class_ = models.Class.objects.get(name="SWISS10")
        tz = pytz.timezone("Europe/Zurich")

        count = models.Count.objects.create(
            start_service_date=tz.localize(datetime(2021, 3, 1)),
            end_service_date=tz.localize(datetime(2021, 3, 2)),
            start_process_date=tz.localize(datetime(2021, 3, 15)),
            end_process_date=tz.localize(datetime(2021, 3, 28)),
            start_put_date=tz.localize(datetime(2021, 2, 20)),
            end_put_date=tz.localize(datetime(2021, 4, 1)),
            id_model=model,
            id_device=device,
            id_sensor_type=sensor_type,
            id_class=class_,
            id_installation=installation,
        )

        for file in Path(test_utils.test_data_path(test_data_folder)).iterdir():
            importer.import_file(test_utils.test_data_path(str(file)), count)

        report.prepare_reports(self.test_outputs, count)
        found_files = len(list(Path(self.test_outputs).iterdir()))
        # The number of files generated is expected to be: weeks measured x sections
        # so let's make sure all sections are considered in the files generation
        self.assertGreater(found_files, 0)
        self.assertEqual(found_files % n_sections, 0)

    def test_all_sections_yearly(self):
        # Test if yearly report features all sections for special case
        test_data_folder = "ASC"
        installation_name = "53309999"

        installation = models.Installation.objects.get(name=installation_name)

        model = models.Model.objects.all().first()
        device = models.Device.objects.all().first()
        sensor_type = models.SensorType.objects.all().first()
        class_ = models.Class.objects.get(name="SWISS10")
        tz = pytz.timezone("Europe/Zurich")

        count = models.Count.objects.create(
            start_put_date=tz.localize(datetime(2021, 1, 1)),
            start_service_date=tz.localize(datetime(2021, 1, 8)),
            end_service_date=tz.localize(datetime(2021, 12, 15)),
            start_process_date=tz.localize(datetime(2021, 1, 15)),
            end_process_date=tz.localize(datetime(2021, 12, 28)),
            end_put_date=tz.localize(datetime(2021, 12, 31)),
            id_model=model,
            id_device=device,
            id_sensor_type=sensor_type,
            id_class=class_,
            id_installation=installation,
        )

        gen = Path(test_utils.test_data_path(test_data_folder)).iterdir()
        to_import = 100
        imported = 0
        for file in islice(gen, to_import):
            importer.import_file(test_utils.test_data_path(str(file)), count)
            imported += 1
            print(f"Remaining: {to_import - imported}")

        sections = models.Section.objects.filter(
            lane__id_installation=installation.id, lane__countdetail__id_count=count.id
        ).distinct()
        self.assertTrue(sections.exists())

        sections_ids = list(sections.values_list("id", flat=True))
        report.prepare_reports(
            file_path=self.test_outputs,
            count=count,
            year=count.start_process_date.year,
            template="yearly",
            sections_ids=sections_ids,
        )
        found_files = len(list(Path(self.test_outputs).iterdir()))
        self.assertEqual(found_files, sections.count())

    def test_yearly_bike_report(self):
        # Import test data pertaining to "mobilité douce"
        installation_name = "64540060"
        installation = models.Installation.objects.get(name=installation_name)
        model = models.Model.objects.all().first()
        device = models.Device.objects.all().first()
        sensor_type = models.SensorType.objects.all().first()
        class_ = models.Class.objects.get(name="SPCH-MD 5C")
        tz = pytz.timezone("Europe/Zurich")

        count = models.Count.objects.create(
            start_service_date=tz.localize(datetime(2021, 2, 1)),
            end_service_date=tz.localize(datetime(2021, 12, 10)),
            start_process_date=tz.localize(datetime(2021, 2, 10)),
            end_process_date=tz.localize(datetime(2021, 12, 15)),
            start_put_date=tz.localize(datetime(2021, 1, 1)),
            end_put_date=tz.localize(datetime(2021, 1, 5)),
            id_model=model,
            id_device=device,
            id_sensor_type=sensor_type,
            id_class=class_,
            id_installation=installation,
        )

        path_to_file = self.test_data.joinpath("64540060_Latenium_PS2021_ChMixte.txt")
        importer.import_file(str(path_to_file), count)
        print("Imported 1 count files!")

        models.CountDetail.objects.update(import_status=0)
        print("Forced import status to 'definitive' for testing purposes")

        sections_ids = (
            models.Section.objects.filter(lane__id_installation__name=installation_name)
            .distinct()
            .values_list("id", flat=True)
        )
        self.assertTrue(sections_ids.exists())

        report = YearlyReportBike(
            path_to_output_dir=self.test_outputs,
            year=2021,
            section_id=sections_ids.first(),
            classtxt="SPCH-MD 5C",
        )
        report.run()

    def test_busiest_by_season(self):
        # Import test data pertaining to "mobilité douce"
        installation_name = "00107695"
        installation = models.Installation.objects.get(name=installation_name)
        model = models.Model.objects.all().first()
        device = models.Device.objects.all().first()
        sensor_type = models.SensorType.objects.all().first()
        class_ = models.Class.objects.get(name="SPCH-MD 5C")
        tz = pytz.timezone("Europe/Zurich")

        count = models.Count.objects.create(
            start_service_date=tz.localize(datetime(2021, 2, 1)),
            end_service_date=tz.localize(datetime(2021, 12, 10)),
            start_process_date=tz.localize(datetime(2021, 2, 10)),
            end_process_date=tz.localize(datetime(2021, 12, 15)),
            start_put_date=tz.localize(datetime(2021, 1, 1)),
            end_put_date=tz.localize(datetime(2021, 1, 5)),
            id_model=model,
            id_device=device,
            id_sensor_type=sensor_type,
            id_class=class_,
            id_installation=installation,
        )

        path_to_file = Path("/test_data").joinpath(
            "64540060_Latenium_PS2021_ChMixte.txt"
        )
        importer.import_file(str(path_to_file), count)
        print("Imported 1 count files!")

        models.CountDetail.objects.update(import_status=0)
        print("Forced import status to 'definitive' for testing purposes")

        sections_ids = (
            models.Section.objects.filter(lane__id_installation__name=installation_name)
            .distinct()
            .values_list("id", flat=True)
        )
        self.assertTrue(sections_ids.exists())

        report = YearlyReportBike(
            path_to_output_dir=self.test_outputs,
            year=2021,
            section_id=sections_ids.first(),
            classtxt="SPCH-MD 5C",
        )
        report.run()

        # Collect count details
        details = report.count_details_by_season(count)
        assert details

        def inspect_leaves(d, res) -> list[int]:
            for v in d.values():
                if isinstance(v, int):
                    res.append(v)
                elif isinstance(v, dict):
                    inspect_leaves(v, res)
            return res

        self.assertTrue(all(value > 0 for value in inspect_leaves(details, [])))

        # Prepare workbook
        path_to_inputs = Path("comptages/report").joinpath("template_yearly_bike.xlsx")
        path_to_outputs = self.test_outputs.joinpath("yearly_bike.xlsx")
        wb = load_workbook(path_to_inputs)

        # Write data & save
        ws = wb["Data_yearly_stats"]
        window = ws["B22:F25"]

        for row, season in zip(window, ("printemps", "été", "automne", "hiver")):
            for cell, category in zip(
                row, ("VELO", "MONO", "SHORT", "SPECIAL", "MULTI")
            ):
                category_data = details.get(season, {}).get(category, [])
                if isinstance(category_data, (list, tuple)):
                    cell.value = sum(details[season][category].values())
                elif isinstance(category_data, int):
                    cell.value = details[season][category]
                else:
                    cell.value = 0

        wb.save(path_to_outputs)

    def test_busiest_by_various_criteria(self):
        # Import test data pertaining to "mobilité douce"
        installation_name = "00107695"
        installation = models.Installation.objects.get(name=installation_name)
        model = models.Model.objects.all().first()
        device = models.Device.objects.all().first()
        sensor_type = models.SensorType.objects.all().first()
        class_ = models.Class.objects.get(name="SPCH-MD 5C")
        tz = pytz.timezone("Europe/Zurich")

        count = models.Count.objects.create(
            start_service_date=tz.localize(datetime(2021, 2, 1)),
            end_service_date=tz.localize(datetime(2021, 12, 10)),
            start_process_date=tz.localize(datetime(2021, 2, 10)),
            end_process_date=tz.localize(datetime(2021, 12, 15)),
            start_put_date=tz.localize(datetime(2021, 1, 1)),
            end_put_date=tz.localize(datetime(2021, 1, 5)),
            id_model=model,
            id_device=device,
            id_sensor_type=sensor_type,
            id_class=class_,
            id_installation=installation,
        )

        path_to_file = Path("/test_data").joinpath(
            "64540060_Latenium_PS2021_ChMixte.txt"
        )
        importer.import_file(str(path_to_file), count)
        print("Imported 1 count files!")

        models.CountDetail.objects.update(import_status=0)
        print("Forced import status to 'definitive' for testing purposes")

        sections_ids = (
            models.Section.objects.filter(lane__id_installation__name=installation_name)
            .distinct()
            .values_list("id", flat=True)
        )
        self.assertTrue(sections_ids.exists())

        report = YearlyReportBike(
            path_to_output_dir=self.test_outputs,
            year=2021,
            section_id=sections_ids.first(),
            classtxt="SPCH-MD 5C",
        )
        report.run()

        # Collecting count details
        data = report.count_details_by_various_criteria(count)

        # Prepare workbook
        path_to_inputs = Path("comptages/report").joinpath("template_yearly_bike.xlsx")
        path_to_outputs = self.test_outputs.joinpath("yearly_bike.xlsx")
        wb = load_workbook(path_to_inputs)

        # Write data & save
        ws = wb["Data_yearly_stats"]
        column_names = (
            "VELO",
            "MONO",
            "SHORT",
            "SPECIAL",
            "MULTI",
            "day_or_month_or_weekend",
        )
        row_names = (
            "total_runs_in_year",
            "busiest_date_row",
            "least_busy_date_row",
            "busiest_month_row",
            "least_busy_month_row",
            "total_runs_busiest_hour_weekday",
            "total_runs_busiest_hour_weekend",
        )
        print_area = ws["B2:G8"]
        for row_idx, row_name in enumerate(row_names, 0):
            row = print_area[row_idx]
            YearlyReportBike.write_to_row(
                row_name=row_name,
                row=row,
                data=data,
                key="category_name",
                column_names=column_names,
            )
        wb.save(path_to_outputs)
