import os
from datetime import date, datetime, timedelta
from typing import Generator, Optional
from qgis.core import Qgis, QgsMessageLog

from openpyxl import load_workbook, Workbook

from comptages.core import statistics
from comptages.datamodel import models


def simple_print_callback(progress):
    print(f"{datetime.now()}: Generating report... {progress}%")


def prepare_reports(
    file_path,
    count: Optional[models.Count] = None,
    year=None,
    template="default",
    sections_ids: Optional[list[str]] = None,
    callback_progress=simple_print_callback,
    sections_days: Optional[dict[str, list[date]]] = None,
):
    print(f"{datetime.now()}: _prepare_reports: begin, folder: {file_path}")
    current_dir = os.path.dirname(os.path.abspath(__file__))

    if template == "default":
        template_name = "template.xlsx"
        template_path = os.path.join(current_dir, os.pardir, "report", template_name)
        assert count
        _prepare_default_reports(
            file_path, count, template_path, callback_progress, sections_days
        )
    elif template == "yearly":
        template_name = "template_yearly.xlsx"
        template_path = os.path.join(current_dir, os.pardir, "report", template_name)
        assert year
        assert sections_ids
        _prepare_yearly_report(
            file_path,
            year,
            template_path,
            sections_ids,
            callback_progress,
            sections_days,
        )
    elif template == "yearly_bike":
        pass
    print(f"{datetime.now()}: _prepare_reports: ended, folder: {file_path}")


def _prepare_default_reports(
    file_path: str,
    count: models.Count,
    template_path: str,
    callback_progress,
    sections_days: Optional[dict[str, list[date]]] = None,
):
    """Write default reports to disk (1 per section in count, per week)"""
    print(f"{datetime.now()}: _prepare_default_reports: begin, count: {count}")
    # We do by section and not by count because of special cases.
    sections = models.Section.objects.filter(
        lane__id_installation__count=count
    ).distinct()
    assert sections
    sections_qty = len(list(sections))

    mondays = list(_mondays_of_count(count))
    assert mondays
    mondays_qty = len(mondays)

    for j, section in enumerate(sections):
        for i, monday in enumerate(mondays):
            # Filter out date based on parameter
            if sections_days and monday not in sections_days[section.id]:
                continue
            output = os.path.join(
                file_path, f"{section.id}_{monday.strftime('%Y%m%d')}_r.xlsx"
            )
            workbook_nbr = i + j * mondays_qty
            QgsMessageLog.logMessage(
                f"{datetime.now()} - Preparing reports: Adding workbook {workbook_nbr} ({output})",
                "Comptages",
                Qgis.Info,
            )
            progress = int(100 / mondays_qty / sections_qty * workbook_nbr)
            callback_progress(progress)

            workbook = load_workbook(filename=template_path)
            _data_count(count, section, monday, workbook)
            _data_day(count, section, monday, workbook)
            _data_speed(count, section, monday, workbook)
            _data_category(count, section, monday, workbook)
            _remove_useless_sheets(count, workbook)

            workbook.save(filename=output)

    print(f"{datetime.now()}: _prepare_default_reports: ended, count: {count}")


def _prepare_yearly_report(
    file_path: str,
    year: int,
    template_path: str,
    sections_ids: list[str],
    callback_progress,
    sections_days: Optional[dict[str, list[date]]] = None,
):
    """Write default reports to disk (1 per section included in the count)"""
    print(
        f"{datetime.now()}: _prepare_yearly_report: begin, sections_ids: {sections_ids}"
    )
    # Get first count to be used as example
    count_qs = models.Count.objects.filter(
        id_installation__lane__id_section=sections_ids[0],
        start_process_date__year__lte=year,
        end_process_date__year__gte=year,
    )
    if not count_qs.exists():
        info_str = f"{datetime.now()}: Aucun comptage trouvé pour cette section {sections_ids[0]} et cette année {year}"
        print(info_str)
        QgsMessageLog.logMessage(info_str, "Comptages", Qgis.Warning)

        return
    count = count_qs.first()
    assert count

    # Filter out sections whose id is not in the sections_ids
    # or whose lanes match no countdetail for the current count
    sections = models.Section.objects.filter(
        id__in=sections_ids,
        lane__countdetail__id_count=count.id,
    ).distinct()

    for section in sections:
        workbook = load_workbook(filename=template_path)
        _data_count_yearly(count, section, year, workbook)
        _data_day_yearly(count, section, year, workbook)
        _data_month_yearly(count, section, year, workbook)
        _data_speed_yearly(count, section, year, workbook)
        _data_category_yearly(count, section, year, workbook)
        _remove_useless_sheets(count, workbook)
        output = os.path.join(file_path, f"{section.id}_{year}_r.xlsx")
        workbook.save(filename=output)

    print(
        f"{datetime.now()}: _prepare_yearly_report: ended, sections_ids: {sections_ids}"
    )


def _mondays_of_count(count: models.Count) -> Generator[date, None, None]:
    """Generator that return the Mondays of the count"""

    start = count.start_process_date
    end = count.end_process_date

    # Monday of first week
    monday = start - timedelta(days=start.weekday())
    yield monday

    while True:
        monday = monday + timedelta(days=7)
        if monday > end:
            return
        yield monday


def _data_count(
    count: models.Count, section: models.Section, monday, workbook: Workbook
):
    ws = workbook["Data_count"]
    ws["B3"] = (
        "Poste de comptage : {}  Axe : {}:{}{}  " "PR {} + {} m à PR {} + {} m"
    ).format(
        section.id,
        section.owner,
        section.road,
        section.way,
        section.start_pr,
        int(section.start_dist),
        section.end_pr,
        int(section.end_dist),
    )

    ws["B4"] = "Periode de comptage du {} au {}".format(
        monday.strftime("%d/%m/%Y"),
        (monday + timedelta(days=7)).strftime("%d/%m/%Y"),
    )

    ws["B5"] = "Comptage {}".format(monday.strftime("%Y"))
    ws["B6"] = "Type de capteur : {}".format(count.id_sensor_type.name)
    ws["B7"] = "Modèle : {}".format(count.id_model.name)
    ws["B8"] = "Classification : {}".format(count.id_class.name)

    ws["B9"] = "Comptage véhicule par véhicule"
    if _is_aggregate(count):
        ws["B9"] = "Comptage par interval"

    special_periods = statistics.get_special_periods(monday, monday + timedelta(days=6))
    texts = []
    for i in special_periods:
        texts.append(f"{i.start_date} - {i.end_date}: {i.description}")
    ws["B10"] = "Periode speciales : {}".format(", ".join(texts))

    ws["B11"] = section.place_name

    if count.remarks:
        ws["B12"] = "Remarque : {}".format(count.remarks)

    lanes = models.Lane.objects.filter(id_section=section).order_by("direction")
    if lanes:
        ws["B13"] = lanes[0].direction_desc
    if len(lanes) > 1:
        ws["B14"] = lanes[1].direction_desc


def _data_count_yearly(
    count: models.Count, section: models.Section, year: int, workbook: Workbook
):
    ws = workbook["Data_count"]
    ws["B3"] = (
        "Poste de comptage : {}  Axe : {}:{}{}  " "PR {} + {} m à PR {} + {} m"
    ).format(
        section.id,
        section.owner,
        section.road,
        section.way,
        section.start_pr,
        int(section.start_dist),
        section.end_pr,
        int(section.end_dist),
    )

    ws["B4"] = "Periode de comptage du 01/01/{0} au 31/12/{0}".format(year)

    ws["B5"] = "Comptage {}".format(year)

    ws["B6"] = "Type de capteur : {}".format(count.id_sensor_type.name)

    ws["B7"] = "Modèle : {}".format(count.id_model.name)

    ws["B8"] = "Classification : {}".format(count.id_class.name)

    ws["B9"] = "Comptage véhicule par véhicule"
    if _is_aggregate(count):
        ws["B9"] = "Comptage par intervale"

    ws["B11"] = section.place_name

    if count.remarks:
        ws["B12"] = "Remarque : {}".format(count.remarks)

    lanes = models.Lane.objects.filter(id_section=section).order_by("direction")
    if lanes:
        ws["B13"] = lanes[0].direction_desc
    if len(lanes) > 1:
        ws["B14"] = lanes[1].direction_desc


def _data_day(
    count: models.Count,
    section: models.Section,
    monday,
    workbook: Workbook,
):
    ws = workbook["Data_day"]

    # Monthly coefficients
    row_offset = 2
    col_offset = 2
    monthly_coefficients = [
        0.93,
        0.96,
        1.00,
        1.02,
        1.01,
        1.04,
        0.98,
        0.98,
        1.04,
        1.03,
        1.02,
        0.98,
    ]

    for i in range(7):
        day = monday + timedelta(days=i)
        ws.cell(
            row=row_offset,
            column=col_offset + i,
            value=monthly_coefficients[day.month - 1],
        )

    # Total (section)
    row_offset = 67
    col_offset = 2
    for i in range(7):
        df = statistics.get_time_data(
            count,
            section,
            start=monday + timedelta(days=i),
            end=monday + timedelta(days=i + 1),
            exclude_trash=True,
        )

        for row in df.itertuples():
            ws.cell(row=row_offset + row.hour, column=col_offset + i, value=row.thm)

    # Direction 1
    row_offset = 5
    col_offset = 2
    for i in range(7):
        df = statistics.get_time_data(
            count,
            section,
            start=monday + timedelta(days=i),
            end=monday + timedelta(days=i + 1),
            direction=1,
            exclude_trash=True,
        )

        for row in df.itertuples():
            ws.cell(row=row_offset + row.hour, column=col_offset + i, value=row.thm)

    # Light heavy direction 1
    row_offset = 30
    col_offset = 2
    for i in range(7):
        light = statistics.get_light_numbers(
            count,
            section,
            start=monday + timedelta(days=i),
            end=monday + timedelta(days=i + 1),
            direction=1,
            exclude_trash=True,
        )
        ws.cell(row=row_offset, column=col_offset + i, value=light.get(True, 0))
        ws.cell(row=row_offset + 1, column=col_offset + i, value=light.get(False, 0))

    # Direction 2
    if len(section.lane_set.all()) == 2:
        row_offset = 36
        col_offset = 2
        for i in range(7):
            df = statistics.get_time_data(
                count,
                section,
                start=monday + timedelta(days=i),
                end=monday + timedelta(days=i + 1),
                direction=2,
                exclude_trash=True,
            )

            for row in df.itertuples():
                ws.cell(row=row_offset + row.hour, column=col_offset + i, value=row.thm)

        # Light heavy direction 2
        row_offset = 61
        col_offset = 2
        for i in range(7):
            light = statistics.get_light_numbers(
                count,
                section,
                start=monday + timedelta(days=i),
                end=monday + timedelta(days=i + 1),
                direction=2,
                exclude_trash=True,
            )
            ws.cell(row=row_offset, column=col_offset + i, value=light.get(True, 0))
            ws.cell(
                row=row_offset + 1, column=col_offset + i, value=light.get(False, 0)
            )


def _data_day_yearly(
    count: models.Count,
    section: models.Section,
    year: int,
    workbook: Workbook,
):
    ws = workbook["Data_day"]

    # Total (section)
    row_offset = 69
    col_offset = 2
    df = statistics.get_time_data_yearly(
        year,
        section,
        exclude_trash=True,
    )

    if df is None:
        print(
            f"{datetime.now()}:_data_day_yearly - Pas de données pour cette section {section} et cette année {year} /!\\/!\\/!\\"
        )
        return

    for i in range(7):
        day_df = df[df["date"] == i]
        for row in day_df.itertuples():
            ws.cell(row=row_offset + row.hour, column=col_offset + i, value=row.thm)

    # Light heavy section
    row_offset = 95
    col_offset = 2
    df = statistics.get_light_numbers_yearly(
        section,
        start=datetime(year, 1, 1),
        end=datetime(year + 1, 1, 1),
        exclude_trash=True,
    )

    for i in range(7):
        ws.cell(
            row=row_offset,
            column=col_offset + i,
            value=int(df[df["date"] == i][df["id_category__light"] == True].value),
        )
        ws.cell(
            row=row_offset + 1,
            column=col_offset + i,
            value=int(df[df["date"] == i][df["id_category__light"] == False].value),
        )

    # Direction 1
    row_offset = 5
    col_offset = 2
    df = statistics.get_time_data_yearly(
        year,
        section,
        direction=1,
        exclude_trash=True,
    )

    if df is None:
        print(
            f"{datetime.now()}:_data_day_yearly - Pas de données pour cette section:{section}, cette direction:1 et cette année:{year} /!\\/!\\/!\\"
        )
        return

    for i in range(7):
        day_df = df[df["date"] == i]
        for row in day_df.itertuples():
            ws.cell(row=row_offset + row.hour, column=col_offset + i, value=row.thm)

    # Light heavy direction 1
    row_offset = 31
    col_offset = 2
    df = statistics.get_light_numbers_yearly(
        section,
        start=datetime(year, 1, 1),
        end=datetime(year + 1, 1, 1),
        direction=1,
        exclude_trash=True,
    )

    for i in range(7):
        ws.cell(
            row=row_offset,
            column=col_offset + i,
            value=int(df[df["date"] == i][df["id_category__light"] == True].value),
        )
        ws.cell(
            row=row_offset + 1,
            column=col_offset + i,
            value=int(df[df["date"] == i][df["id_category__light"] == False].value),
        )

    # Direction 2
    if len(section.lane_set.all()) == 2:
        row_offset = 37
        col_offset = 2
        df = statistics.get_time_data_yearly(
            year,
            section,
            direction=2,
            exclude_trash=True,
        )

        for i in range(7):
            day_df = df[df["date"] == i]
            for row in day_df.itertuples():
                ws.cell(row=row_offset + row.hour, column=col_offset + i, value=row.thm)

        # Light heavy direction 2
        row_offset = 63
        col_offset = 2
        df = statistics.get_light_numbers_yearly(
            section,
            start=datetime(year, 1, 1),
            end=datetime(year + 1, 1, 1),
            direction=2,
            exclude_trash=True,
        )

        for i in range(7):
            ws.cell(
                row=row_offset,
                column=col_offset + i,
                value=int(df[df["date"] == i][df["id_category__light"] == True].value),
            )
            ws.cell(
                row=row_offset + 1,
                column=col_offset + i,
                value=int(df[df["date"] == i][df["id_category__light"] == False].value),
            )


def _data_month_yearly(
    count: models.Count, section: models.Section, year: int, workbook: Workbook
):
    ws = workbook["Data_month"]
    start = datetime(year, 1, 1)
    end = datetime(year + 1, 1, 1)

    # Section
    row_offset = 14
    col_offset = 2
    df = statistics.get_month_data(
        section,
        start,
        end,
        exclude_trash=True,
    )

    for col in df.itertuples():
        ws.cell(row=row_offset, column=col_offset + col.Index, value=col.tm)

    # Direction 1
    row_offset = 4
    col_offset = 2
    df = statistics.get_month_data(
        section,
        start,
        end,
        direction=1,
        exclude_trash=True,
    )

    for col in df.itertuples():
        ws.cell(row=row_offset, column=col_offset + col.Index, value=col.tm)

    # Direction 2
    row_offset = 9
    col_offset = 2
    df = statistics.get_month_data(
        section,
        start,
        end,
        direction=2,
        exclude_trash=True,
    )

    for col in df.itertuples():
        ws.cell(row=row_offset, column=col_offset + col.Index, value=col.tm)

    # Monthly coefficients
    row_offset = 18
    col_offset = 2
    monthly_coefficients = [
        0.93,
        0.96,
        1.00,
        1.02,
        1.01,
        1.04,
        0.98,
        0.98,
        1.04,
        1.03,
        1.02,
        0.98,
    ]

    for i in range(12):
        ws.cell(
            row=row_offset,
            column=col_offset + i,
            # FIXME: calculate actual coefficients
            value=monthly_coefficients[i],
        )


def _data_speed(
    count: models.Count, section: models.Section, monday, workbook: Workbook
):
    ws = workbook["Data_speed"]

    speed_ranges = [
        (0, 10),
        (10, 20),
        (20, 30),
        (30, 40),
        (40, 50),
        (50, 60),
        (60, 70),
        (70, 80),
        (80, 90),
        (90, 100),
        (100, 110),
        (110, 120),
        (120, 999),
    ]

    if _is_aggregate(count):
        speed_ranges = [
            (0, 30),
            (30, 40),
            (40, 50),
            (50, 60),
            (60, 70),
            (70, 80),
            (80, 90),
            (90, 100),
            (100, 110),
            (110, 120),
            (120, 130),
            (130, 999),
        ]

    characteristic_speeds = [0.15, 0.5, 0.85]

    # Direction 1
    row_offset = 5
    col_offset = 2
    for i, range_ in enumerate(speed_ranges):
        res = statistics.get_speed_data_by_hour(
            count,
            section,
            direction=1,
            start=monday,
            end=monday + timedelta(days=7),
            speed_low=range_[0],
            speed_high=range_[1],
            exclude_trash=True,
        )
        for row in res:
            ws.cell(row=row_offset + row[0], column=col_offset + i, value=row[1])

    if not _is_aggregate(count):
        # Characteristic speed direction 1
        row_offset = 5
        col_offset = 16
        for i, v in enumerate(characteristic_speeds):
            df = statistics.get_characteristic_speed_by_hour(
                count,
                section,
                direction=1,
                start=monday,
                end=monday + timedelta(days=7),
                v=v,
                exclude_trash=True,
            )
            for row in df.itertuples():
                ws.cell(
                    row=row_offset + row.Index, column=col_offset + i, value=row.speed
                )

        # Average speed direction 1
        row_offset = 5
        col_offset = 19
        df = statistics.get_average_speed_by_hour(
            count,
            section,
            direction=1,
            start=monday,
            end=monday + timedelta(days=7),
            exclude_trash=True,
        )
        for row in df.itertuples():
            ws.cell(row=row_offset + row.Index, column=col_offset, value=row.speed)

    # Direction 2
    if len(section.lane_set.all()) == 2:
        row_offset = 33
        col_offset = 2
        for i, range_ in enumerate(speed_ranges):
            res = statistics.get_speed_data_by_hour(
                count,
                section,
                direction=2,
                start=monday,
                end=monday + timedelta(days=7),
                speed_low=range_[0],
                speed_high=range_[1],
                exclude_trash=True,
            )
            for row in res:
                ws.cell(row=row_offset + row[0], column=col_offset + i, value=row[1])

        if not _is_aggregate(count):
            # Characteristic speed direction 2
            row_offset = 33
            col_offset = 16
            for i, v in enumerate(characteristic_speeds):
                df = statistics.get_characteristic_speed_by_hour(
                    count,
                    section,
                    direction=2,
                    start=monday,
                    end=monday + timedelta(days=7),
                    v=v,
                    exclude_trash=True,
                )
                for row in df.itertuples():
                    ws.cell(
                        row=row_offset + row.Index,
                        column=col_offset + i,
                        value=row.speed,
                    )

        # Average speed direction 2
        row_offset = 33
        col_offset = 19
        df = statistics.get_average_speed_by_hour(
            count,
            section,
            direction=2,
            start=monday,
            end=monday + timedelta(days=7),
            exclude_trash=True,
        )
        for row in df.itertuples():
            ws.cell(row=row_offset + row.Index, column=col_offset, value=row.speed)


def _data_speed_yearly(
    count: models.Count, section: models.Section, year: int, workbook: Workbook
):
    ws = workbook["Data_speed"]
    start = datetime(year, 1, 1)
    end = datetime(year + 1, 1, 1)

    speed_ranges = [
        (0, 10),
        (10, 20),
        (20, 30),
        (30, 40),
        (40, 50),
        (50, 60),
        (60, 70),
        (70, 80),
        (80, 90),
        (90, 100),
        (100, 110),
        (110, 120),
        (120, 999),
    ]

    if _is_aggregate(count):
        speed_ranges = [
            (0, 30),
            (30, 40),
            (40, 50),
            (50, 60),
            (60, 70),
            (70, 80),
            (80, 90),
            (90, 100),
            (100, 110),
            (110, 120),
            (120, 130),
            (130, 999),
        ]

    characteristic_speeds = [0.15, 0.5, 0.85]

    # Direction 1
    row_offset = 5
    col_offset = 2
    for i, range_ in enumerate(speed_ranges):
        res = statistics.get_speed_data_by_hour(
            None,
            section,
            direction=1,
            start=start,
            end=end,
            speed_low=range_[0],
            speed_high=range_[1],
            exclude_trash=True,
        )
        for row in res:
            ws.cell(row=row_offset + row[0], column=col_offset + i, value=row[1])

    if not _is_aggregate(count):
        # Characteristic speed direction 1
        row_offset = 5
        col_offset = 16
        for i, v in enumerate(characteristic_speeds):
            df = statistics.get_characteristic_speed_by_hour(
                None,
                section,
                direction=1,
                start=start,
                end=end,
                v=v,
                exclude_trash=True,
            )
            for row in df.itertuples():
                ws.cell(
                    row=row_offset + row.Index, column=col_offset + i, value=row.speed
                )

        # Average speed direction 1
        row_offset = 5
        col_offset = 19
        df = statistics.get_average_speed_by_hour(
            None,
            section,
            direction=1,
            start=start,
            end=end,
            exclude_trash=True,
        )
        for row in df.itertuples():
            ws.cell(row=row_offset + row.Index, column=col_offset, value=row.speed)

    # Direction 2
    if len(section.lane_set.all()) == 2:
        row_offset = 33
        col_offset = 2
        for i, range_ in enumerate(speed_ranges):
            res = statistics.get_speed_data_by_hour(
                None,
                section,
                direction=2,
                start=start,
                end=end,
                speed_low=range_[0],
                speed_high=range_[1],
                exclude_trash=True,
            )

            for row in res:
                ws.cell(row=row_offset + row[0], column=col_offset + i, value=row[1])

        if not _is_aggregate(count):
            # Characteristic speed direction 2
            row_offset = 33
            col_offset = 16
            for i, v in enumerate(characteristic_speeds):
                df = statistics.get_characteristic_speed_by_hour(
                    None,
                    section,
                    direction=2,
                    start=start,
                    end=end,
                    v=v,
                    exclude_trash=True,
                )
                for row in df.itertuples():
                    ws.cell(
                        row=row_offset + row.Index,
                        column=col_offset + i,
                        value=row.speed,
                    )

            # Average speed direction 2
            row_offset = 33
            col_offset = 19
            df = statistics.get_average_speed_by_hour(
                None,
                section,
                direction=2,
                start=start,
                end=end,
                exclude_trash=True,
            )
            for row in df.itertuples():
                ws.cell(row=row_offset + row.Index, column=col_offset, value=row.speed)


def _data_category(
    count: models.Count, section: models.Section, monday, workbook: Workbook
):
    ws = workbook["Data_category"]

    categories = (
        models.Category.objects.filter(countdetail__id_count=count)
        .distinct()
        .order_by("code")
    )

    # Direction 1
    row_offset = 5
    col_offset = 2
    for category in categories:
        res = statistics.get_category_data_by_hour(
            count,
            section,
            category=category,
            direction=1,
            start=monday,
            end=monday + timedelta(days=7),
        )

        for row in res:
            row_num = row_offset + row[0]
            col_num = col_offset + _t_cat(count, category.code)
            value = (
                ws.cell(row_num, col_num).value + row[1]
            )  # Add to previous value because with class convertions multiple categories can converge into a single one

            ws.cell(row=row_num, column=col_num, value=value)

    # Direction 2
    if len(section.lane_set.all()) == 2:
        row_offset = 33
        col_offset = 2
        for category in categories:
            res = statistics.get_category_data_by_hour(
                count,
                section,
                category=category,
                direction=2,
                start=monday,
                end=monday + timedelta(days=7),
            )

            for row in res:
                row_num = row_offset + row[0]
                col_num = col_offset + _t_cat(count, category.code)
                value = (
                    ws.cell(row_num, col_num).value + row[1]
                )  # Add to previous value because with class convertions multiple categories can converge into a single one

                ws.cell(row=row_num, column=col_num, value=value)


def _data_category_yearly(
    count: models.Count, section: models.Section, year: int, workbook: Workbook
):
    ws = workbook["Data_category"]
    start = datetime(year, 1, 1)
    end = datetime(year + 1, 1, 1)

    categories = (
        models.Category.objects.filter(countdetail__id_count=count)
        .distinct()
        .order_by("code")
    )

    # Direction 1
    row_offset = 5
    col_offset = 2
    for category in categories:
        res = statistics.get_category_data_by_hour(
            None,
            section,
            category=category,
            direction=1,
            start=start,
            end=end,
        )

        for row in res:
            row_num = row_offset + row[0]
            col_num = col_offset + _t_cat(count, category.code)
            value = (
                ws.cell(row_num, col_num).value + row[1]
            )  # Add to previous value because with class convertions multiple categories can converge into a single one

            ws.cell(row=row_num, column=col_num, value=value)

    # Direction 2
    if len(section.lane_set.all()) == 2:
        row_offset = 33
        col_offset = 2
        for category in categories:
            res = statistics.get_category_data_by_hour(
                None,
                section,
                category=category,
                direction=2,
                start=start,
                end=end,
            )

            for row in res:
                row_num = row_offset + row[0]
                col_num = col_offset + _t_cat(count, category.code)
                value = (
                    ws.cell(row_num, col_num).value + row[1]
                )  # Add to previous value because with class convertions multiple categories can converge into a single one

                ws.cell(row=row_num, column=col_num, value=value)


def _remove_useless_sheets(count: models.Count, workbook: Workbook):
    class_name = _t_cl(count.id_class.name)

    if class_name == "SWISS10":
        to_remove_from_spreadsheet = [
            "SWISS7_H",
            "SWISS7_G",
            "EUR6_H",
            "EUR6_G",
        ]
    elif class_name == "SWISS7":
        to_remove_from_spreadsheet = ["SWISS10_H", "SWISS10_G", "EUR6_H", "EUR6_G"]
    elif class_name == "EUR6":
        to_remove_from_spreadsheet = [
            "SWISS10_H",
            "SWISS10_G",
            "SWISS7_H",
            "SWISS7_G",
        ]
    elif class_name == "Volume1":
        to_remove_from_spreadsheet = [
            "SWISS10_H",
            "SWISS10_G",
            "SWISS7_H",
            "SWISS7_G",
            "EUR6_H",
            "EUR6_G",
        ]
    else:
        to_remove_from_spreadsheet = []

    if _is_aggregate(count):
        to_remove_from_spreadsheet.append("Vit_Hd")
    else:
        to_remove_from_spreadsheet.append("Vit_H")

    for key in to_remove_from_spreadsheet:
        try:
            workbook.remove(workbook[key])
        except KeyError:
            continue


def _t_cl(class_name):
    """Translate class name"""

    if class_name == "FHWA13":
        return "SWISS7"

    if class_name is None:
        return "Volume1"

    if class_name == "SPCH13":
        return "SWISS7"

    return class_name


def _t_cat(count: models.Count, cat_id):
    """Convert categories of a class into the ones of another class e.g.
    FHWA13 should be converted in SWISS7 in order to fill the
    report cells
    """

    if count.id_class.name == "ARXCycle13":
        # FIXME: implement real conversiont between ARX Cycle and SWISS7 or 10
        new_hour = [0] * 7
        return new_hour

    if count.id_class.name == "FHWA13":
        conv = {
            0: 0,
            1: 2,
            2: 3,
            3: 3,
            4: 4,
            5: 4,
            6: 4,
            7: 1,
            8: 5,
            9: 7,
            10: 6,
            11: 7,
            12: 7,
            13: 7,
            14: 7,
        }
        return conv[cat_id]

    if count.id_class.name == "SPCH13":
        conv = {
            0: 0,
            1: 2,
            2: 3,
            3: 3,
            4: 4,
            5: 4,
            6: 4,
            7: 1,
            8: 5,
            9: 6,
            10: 7,
            11: 6,
            12: 6,
            13: 7,
        }
        return conv[cat_id]

    if count.id_class.name == "EUR6":
        conv = {
            0: 0,
            1: 2,
            2: 3,
            3: 4,
            4: 5,
            5: 6,
            6: 1,
        }
        return conv[cat_id]

    return cat_id if cat_id < 11 else 10


def _is_aggregate(count):
    from_aggregate = (
        models.CountDetail.objects.filter(id_count=count)
        .distinct("from_aggregate")
        .values("from_aggregate")
    )
    assert from_aggregate.exists()
    return from_aggregate[0]["from_aggregate"]
