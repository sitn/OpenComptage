from datetime import datetime, timedelta
from functools import reduce
from os import path
from typing import Any, Iterable, Optional, Union
from decimal import Decimal
from qgis.core import Qgis, QgsMessageLog

from openpyxl import load_workbook
from django.db.models import Sum, F, Avg
from django.db.models.functions import (
    ExtractHour,
    ExtractIsoWeekDay,
    ExtractMonth,
    TruncDate,
)

from comptages.core import definitions, utils
from comptages.datamodel.models import (
    CountDetail,
    Section,
    Lane,
    ClassCategory,
    Category,
    Count as modelCount,
)


class YearlyReportBike:
    def __init__(self, path_to_output_dir, year, section_id, classtxt):
        # TODO: pass section or section id?

        self.path_to_output_dir = path_to_output_dir
        self.year = year
        self.section_id = section_id
        self.classtxt = classtxt
        self.seasons = {
            "printemps": [3, 4, 5],
            "été": [6, 7, 8],
            "automne": [9, 10, 11],
            "hiver": [12, 1, 2],
        }

    def total_runs_by_directions(self) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            # id_category__code__in=[1, 2],
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # Total by day of the week (0->monday, 7->sunday) and by direction
        return (
            qs.annotate(weekday=ExtractIsoWeekDay("timestamp"))
            .values("weekday")
            .annotate(total=Sum("times"))
            .values("weekday", "id_lane__direction", "total")
        )

    def tjms_by_weekday_category(
        self, direction=None
    ) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        # Test/GL
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        if direction is not None:
            qs = qs.filter(id_lane__direction=direction)

        # Total by day of the week (1->monday, 7->sunday) and by category (0->14)
        results = (
            qs.annotate(weekday=ExtractIsoWeekDay("timestamp"))
            .values("weekday")
            .annotate(tjm=Sum("times"))
            .values("weekday", "id_category__code", "tjm")
        )
        print(
            f"yearly_report_bike.py: tjms_by_weekday_category - results.query:{str(results.query)}"
        )
        return results

    def tjms_by_weekday_hour(self) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        # Doesn't produce correct results because of 2 aggregation
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # TODO: don't divide by 51 but actually aggregate first by the
        # real days (with sum) and then aggregate by weekday (with average)

        # Total by day of the week (0->monday, 6->sunday) and by hour (0->23)
        results = (
            qs.annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(tj=Sum("times"))
            .values("date", "tj")
            .annotate(weekday=ExtractIsoWeekDay("date"), hour=ExtractHour("timestamp"))
            .values("weekday", "hour")
            .annotate(tjm=Avg("tj"))
            .values("weekday", "hour", "tjm")
        )
        print(
            f"yearly_report_bike.py : tjms_by_weekday_hour - results.query:{str(results.query)}"
        )
        return results

    def total_runs_by_hour_and_direction(
        self, directions=(1, 2), weekdays=(1, 2, 3, 4, 5, 6, 7)
    ) -> dict[int, Any]:
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            id_lane__direction__in=directions,
            timestamp__iso_week_day__in=weekdays,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        results = (
            qs.annotate(hour=ExtractHour("timestamp"))
            .values("hour")
            .annotate(
                runs=Sum("times"),
                direction=F("id_lane__direction"),
                section=F("id_lane__id_section_id"),
            )
            .values("runs", "hour", "direction", "section")
        )
        print(
            f"yearly_report_bike.py : total_runs_by_hour_and_direction - results.query:{str(results.query)}"
        )

        def partition(acc: dict, val: dict) -> dict:
            hour = val["hour"]
            direction = val["direction"]

            if hour not in acc:
                acc[hour] = {}

            if direction not in acc[hour]:
                acc[hour][direction] = {}

            acc[hour][direction] = val
            return acc

        return reduce(partition, results, {})

    def total_runs_by_hour_one_direction(self, direction: int) -> dict[int, Any]:
        # Get all the count details for hours and the specific direction
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            id_lane__direction=direction,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        results = (
            qs.annotate(hour=ExtractHour("timestamp"))
            .values("hour")
            .annotate(runs=Sum("times"))
            .values("runs", "hour")
            .annotate(day=ExtractIsoWeekDay("timestamp"))
            .order_by("day")
        )
        print(
            f"yearly_report_bike.py : total_runs_by_hour_one_direction - results.query:{str(results.query)}"
        )

        def reducer(acc: dict, val: dict) -> dict:
            day = val["day"]
            hour = val["hour"]

            if day not in acc:
                acc[day] = {}

            if hour not in acc[day]:
                acc[day][hour] = {}

            acc[day][hour] = val["runs"]
            return acc

        return reduce(reducer, results, {})

    def total_runs_by_hour_weekday_one_direction(
        self, direction: int
    ) -> "ValuesQuerySet[Countdetail, dict[str, Any]]":
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            id_lane__direction=direction,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        results = (
            qs.annotate(hour=ExtractHour("timestamp"))
            .values("hour")
            .annotate(runs=Sum("times"))
            .values("runs", "hour")
            .annotate(day=ExtractIsoWeekDay("timestamp"))
            .order_by("day")
        )
        print(
            f"yearly_report_bike.py : total_runs_by_hour_weekday_one_direction - results.query:{str(results.query)}"
        )
        return results

    def tjms_by_weekday_and_month(
        self,
    ) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # TODO: don't divide by 12 but actually aggregate first by the
        # real days (with sum) and then aggregate by weekday (with average)

        # Total by day of the week (0->monday, 6->sunday) and by month (1->12)
        results = (
            qs.annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(daily_runs=Sum("times"))
            .values("date", "daily_runs")
            .annotate(week_day=ExtractIsoWeekDay("timestamp"))
            .values("date", "daily_runs", "week_day")
            .annotate(month=ExtractMonth("timestamp"))
            .values("week_day", "month", "daily_runs")
        )
        print(
            f"yearly_report_bike.py : tjms_by_weekday_and_month - results.query:{str(results.query)}"
        )

        # FIXME
        # Aggregation via `values()` into `annotate()` all the way to the end result would be more performant.
        builder = {}
        for item in results:
            if item["month"] not in builder:
                builder[item["month"]] = {"month": item["month"]}
                builder[item["month"]][item["week_day"]] = {
                    "days": 1,
                    "runs": item["daily_runs"],
                    "tjm": 0,
                    "week_day": item["week_day"],
                }
            elif item["week_day"] not in builder[item["month"]]:
                builder[item["month"]][item["week_day"]] = {
                    "days": 1,
                    "runs": item["daily_runs"],
                    "tjm": 0,
                    "week_day": item["week_day"],
                }
            else:
                builder[item["month"]][item["week_day"]]["days"] += 1
                builder[item["month"]][item["week_day"]]["runs"] += item["daily_runs"]
                builder[item["month"]][item["week_day"]]["tjm"] = (
                    builder[item["month"]][item["week_day"]]["runs"]
                    / builder[item["month"]][item["week_day"]]["days"]
                )

        return builder

    def runs_by_weekday_and_month(
        self,
    ) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # Group by month, week_day
        results = (
            qs.annotate(
                month=ExtractMonth("timestamp"), week_day=ExtractIsoWeekDay("timestamp")
            )
            .values("month", "week_day")
            .annotate(daily_runs=Sum("times"))
            .values("month", "week_day", "daily_runs")
        )
        print(
            f"yearly_report_bike.py : runs_by_weekday_and_month - results.query:{str(results.query)}"
        )

        return results

    def nb_weekday_by_month(self) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # Group by date then by month, week_day
        results = (
            qs.annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(daily_runs=Sum("times"))
            .values("date", "daily_runs")
            .annotate(
                month=ExtractMonth("timestamp"), week_day=ExtractIsoWeekDay("timestamp")
            )
            .values("date", "month", "week_day")
            # .order_by("date")
        )
        print(
            f"yearly_report_bike.py : nb_weekday_by_month - results.query:{str(results.query)}"
        )

        def reducer(acc: dict, item) -> dict:

            if item["month"] not in acc:
                acc[item["month"]] = {}

            if item["week_day"] not in acc[item["month"]]:
                acc[item["month"]][item["week_day"]] = 0

            acc[item["month"]][item["week_day"]] += 1

            return acc

        # Collecting
        return reduce(reducer, results, {})

    def total_runs_by_day(self) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # Group by date
        results = (
            qs.annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(daily_runs=Sum("times"))
            .values("date", "daily_runs")
            # .order_by("date")
        )
        print(
            f"yearly_report_bike.py : total_runs_by_day - results.query:{str(results.query)}"
        )

        return results

    def tjms_total_runs_by_day_of_week(self) -> dict[str, Any]:
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        results = (
            qs.annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(daily_runs=Sum("times"), week_day=ExtractIsoWeekDay("timestamp"))
            .values("week_day", "daily_runs")
            .order_by("week_day")
        )
        print(
            f"yearly_report_bike.py : tjms_total_runs_by_day_of_week - results.query:{str(results.query)}"
        )

        # FIXME
        # Aggregation via `values()` into `annotate()` all the way to the end result would be more performant.
        builder = {}
        for item in results:
            if item["week_day"] not in builder:
                builder[item["week_day"]] = {
                    "days": 1,
                    "runs": item["daily_runs"],
                    "tjm": item["daily_runs"],
                    "week_day": item["week_day"],
                }
            else:
                builder[item["week_day"]]["days"] += 1
                builder[item["week_day"]]["runs"] += item["daily_runs"]
                builder[item["week_day"]]["tjm"] = (
                    builder[item["week_day"]]["runs"]
                    / builder[item["week_day"]]["days"]
                )

        return builder

    def total_runs_by_class(self) -> dict[str, Any]:
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        results = (
            qs.annotate(day=ExtractIsoWeekDay("timestamp"))
            .values("day")
            .annotate(runs=Sum("times"), code=F("id_category__code"))
            .values("day", "runs", "code")
        )
        print(
            f"yearly_report_bike.py : total_runs_by_class - results.query:{str(results.query)}"
        )

        def reducer(acc: dict, i: dict):
            code = i["code"]
            day = i["day"]
            runs = i["runs"]

            if code not in acc:
                acc[code] = {}

            acc[code][day] = runs
            return acc

        return reduce(reducer, results, {})

    def tjms_by_direction_bike(
        self, categories, direction, weekdays=[1, 2, 3, 4, 5, 6, 7]
    ) -> float:
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            timestamp__iso_week_day__in=weekdays,
            id_category__code__in=categories,
            id_lane__direction=direction,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        assert qs.exists()
        results = qs.aggregate(res=Sum("times"))["res"]
        print(
            f"yearly_report_bike.py : tjms_by_direction_bike - results.query:{str(results.query)}"
        )
        # TODO: avoid the division?
        return results / 365

    def total(self, categories=[1]) -> float:
        qs = CountDetail.objects.filter(
            timestamp__year=self.year,
            id_category__code__in=categories,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        results = qs.aggregate(res=Sum("times"))["res"]
        print(f"yearly_report_bike.py : total - results.query:{str(results.query)}")

        return results

    def max_day(self, categories=[1]) -> tuple[str, Any]:
        qs = (
            CountDetail.objects.filter(
                timestamp__year=self.year,
                id_category__code__in=categories,
                import_status=definitions.IMPORT_STATUS_DEFINITIVE,
            )
            .annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(total=Sum("times"))
            .order_by("-total")
        )
        print(f"yearly_report_bike.py : max_day - qs.query:{str(qs.query)}")

        return qs[0]["total"], qs[0]["date"]

    def max_month(self, categories=[1]) -> tuple[str, Any]:
        qs = (
            CountDetail.objects.filter(
                timestamp__year=self.year,
                id_category__code__in=categories,
                import_status=definitions.IMPORT_STATUS_DEFINITIVE,
            )
            .annotate(month=ExtractMonth("timestamp"))
            .values("month")
            .annotate(total=Sum("times"))
            .order_by("-total")
        )
        print(f"yearly_report_bike.py : max_month - qs.query:{str(qs.query)}")

        return qs[0]["total"], qs[0]["month"]

    def min_month(self, categories=[1]) -> tuple[str, Any]:
        qs = (
            CountDetail.objects.filter(
                timestamp__year=self.year,
                id_category__code__in=categories,
                import_status=definitions.IMPORT_STATUS_DEFINITIVE,
            )
            .annotate(month=ExtractMonth("timestamp"))
            .values("month")
            .annotate(total=Sum("times"))
            .order_by("total")
        )
        print(f"yearly_report_bike.py : min_month - qs.query:{str(qs.query)}")

        return qs[0]["total"], qs[0]["month"]

    @staticmethod
    def count_details_by_day_month(self, count: modelCount) -> dict[int, Any]:
        # Preparing to filter out categories that don't reference the class picked out by `class_name`
        class_name = self.classtxt
        # Excluding irrelevant
        categories_name_to_exclude = ("TRASH", "ELSE")
        categories_ids = (
            ClassCategory.objects.filter(id_class__name=class_name)
            .exclude(id_category__name__in=categories_name_to_exclude)
            .values_list("id_category", flat=True)
        )
        qs = (
            CountDetail.objects.filter(
                id_count=count.id, id_category__in=categories_ids
            )
            .annotate(
                month=ExtractMonth("timestamp"), day=ExtractIsoWeekDay("timestamp")
            )
            .values("month", "day")
            .annotate(Sum("times"))
        )
        print(
            f"yearly_report_bike.py : count_details_by_day_month - qs.query:{str(qs.query)}"
        )

        def reducer(acc, item):
            month = item["month"]
            day = item["day"]
            runs = item["times__sum"]
            if month not in acc:
                acc[month] = {}
            if day not in acc[month]:
                acc[month][day] = runs
            return acc

        return reduce(reducer, qs, {})

    @staticmethod
    def count_details_by_various_criteria(
        self,
        count: modelCount,
    ) -> dict[str, tuple["ValueQuerySet[CountDetail]", Optional[str]]]:
        # Preparing to filter out categories that don't reference the class picked out by `class_name`
        class_name = self.classtxt
        # Excluding irrelevant
        categories_name_to_exclude = ("TRASH", "ELSE")
        categories_ids = (
            ClassCategory.objects.filter(id_class__name=class_name)
            .exclude(id_category__name__in=categories_name_to_exclude)
            .values_list("id_category", flat=True)
        )
        print(
            f"yearly_report_bike.py : count_details_by_various_criteria - categories_ids.query:{str(categories_ids.query)}"
        )

        # Base QuerySet
        base_qs = CountDetail.objects.filter(
            id_count=count.id,
            id_category__in=categories_ids,
            timestamp__year=self.year,
        )
        print(
            f"yearly_report_bike.py : count_details_by_various_criteria - base_qs.query:{str(base_qs.query)}"
        )

        # Specialized QuerySets
        total_runs_in_year = (
            base_qs.annotate(category_name=F("id_category__name"))
            .values("category_name")
            .annotate(value=Sum("times"))
        )
        print(
            f"yearly_report_bike.py : count_details_by_various_criteria - total_runs_in_year.query:{str(total_runs_in_year.query)}"
        )

        busy_date = (
            base_qs.annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(value=Sum("times"))
            .order_by("-value")
        )
        print(
            f"yearly_report_bike.py : count_details_by_various_criteria - busy_date.query:{str(busy_date.query)}"
        )

        busiest_date = busy_date.first()
        least_busy_date = busy_date.last()
        assert busiest_date
        assert least_busy_date

        busiest_date_row = (
            base_qs.annotate(
                date=TruncDate("timestamp"), category_name=F("id_category__name")
            )
            .filter(date=busiest_date["date"])
            .values("date", "category_name")
            .annotate(value=Sum("times"))
        )
        print(
            f"yearly_report_bike.py : count_details_by_various_criteria - busiest_date_row.query:{str(busiest_date_row.query)}"
        )

        least_busy_date_row = (
            base_qs.annotate(
                date=TruncDate("timestamp"), category_name=F("id_category__name")
            )
            .filter(date=least_busy_date["date"])
            .values("date", "category_name")
            .annotate(value=Sum("times"))
        )
        print(
            f"yearly_report_bike.py : count_details_by_various_criteria - least_busy_date_row.query:{str(least_busy_date_row.query)}"
        )

        busy_month = (
            base_qs.annotate(month=ExtractMonth("timestamp"))
            .values("month")
            .annotate(value=Sum("times"))
            .order_by("-value")
        )
        print(
            f"yearly_report_bike.py : count_details_by_various_criteria - busy_month.query:{str(busy_month.query)}"
        )

        busiest_month = busy_month.first()
        least_busy_month = busy_month.last()
        assert busiest_month
        assert least_busy_month

        busiest_month_row = (
            base_qs.annotate(
                month=ExtractMonth("timestamp"), category_name=F("id_category__name")
            )
            .filter(month=busiest_month["month"])
            .values("month", "category_name")
            .annotate(value=Sum("times"))
        )
        least_busy_month_row = (
            base_qs.annotate(
                month=ExtractMonth("timestamp"), category_name=F("id_category__name")
            )
            .filter(month=least_busy_month["month"])
            .values("month", "category_name")
            .annotate(value=Sum("times"))
        )

        busiest_hour = (
            base_qs.annotate(
                category_name=F("id_category__name"),
                date=TruncDate("timestamp"),
                hour=ExtractHour("timestamp"),
                week_day=ExtractIsoWeekDay("timestamp"),
            )
            .values("date", "hour", "category_name")
            .annotate(value=Sum("times"))
            .order_by("-value")
        )

        total_runs_busiest_hour_weekday = busiest_hour.exclude(week_day__gt=5)
        total_runs_busiest_hour_weekend = busiest_hour.exclude(week_day__lt=6)
        print(
            f"yearly_report_bike.py : count_details_by_various_criteria - busiest_weekend_hour.query:{str(total_runs_busiest_hour_weekend.query)}"
        )

        busiest_weekday = total_runs_busiest_hour_weekday.first()
        busiest_weekend = total_runs_busiest_hour_weekend.first()
        assert busiest_weekday
        assert busiest_weekend

        return {
            "busiest_date_row": (busiest_date_row, busiest_date["date"]),
            "least_busy_date_row": (least_busy_date_row, str(least_busy_date["date"])),
            "busiest_month_row": (busiest_month_row, str(busiest_month["month"])),
            "least_busy_month_row": (
                least_busy_month_row,
                str(least_busy_month["month"]),
            ),
            "total_runs_busiest_hour_weekday": (
                total_runs_busiest_hour_weekday,
                f'{busiest_weekday["date"]} {busiest_weekday["hour"]}:00',
            ),
            "total_runs_busiest_hour_weekend": (
                total_runs_busiest_hour_weekend,
                f'{busiest_weekend["date"]} {busiest_weekend["hour"]}:00',
            ),
            "total_runs_in_year": (total_runs_in_year, None),
        }

    @staticmethod
    def count_details_by_season(self, count_id) -> dict[int, Any]:
        """Break down count details by season x section x class"""
        # Assuming seasons to run from 21 <month> to 20 <month n + 3> -> month20 = (date - timedelta(days=20)).month
        # Preparing to filter out categories that don't reference the class picked out by `class_name`
        class_name = self.classtxt
        # Excluding irrelevant
        categories_name_to_exclude = ("TRASH", "ELSE")
        categories_ids = (
            ClassCategory.objects.filter(id_class__name=class_name)
            .exclude(id_category__name__in=categories_name_to_exclude)
            .values_list("id_category", flat=True)
        )
        # Getting data
        count_details = (
            CountDetail.objects.filter(
                id_count=count_id,
                id_category__in=categories_ids,
                timestamp__year=self.year,
            )
            .annotate(
                date=TruncDate("timestamp"),
                category_name=F("id_category__name"),
            )
            .values("date", "category_name")
            .annotate(value=Sum("times"))
            .values("date", "category_name", "value")
        )
        print(
            f"yearly_report_bike.py : count_details_by_season - count_details.query:{str(count_details.query)}"
        )

        # Preparing to collect data
        def reducer(acc: dict, detail) -> dict:
            date: datetime = detail["date"]
            month20 = (date - timedelta(days=20)).month

            for season, _range in self.seasons.items():
                if month20 in _range:
                    category_name = detail["category_name"]

                    if season not in acc:
                        acc[season] = {}

                    if category_name not in acc[season]:
                        acc[season][category_name] = 0

                    acc[season][category_name] += detail["value"]
                    break

            return acc

        # Collecting
        return reduce(reducer, count_details, {})

    @staticmethod
    def write_to_row(
        *,
        row_name: str,
        row: Iterable,
        data: dict,
        key: str,
        column_names: Iterable[str],
    ):
        items, day_or_month_or_weekend = data[row_name]
        for column_name, cell in zip(column_names, row):
            if column_name == "day_or_month_or_weekend":
                cell.value = day_or_month_or_weekend or "-"
            elif item := next(
                filter(
                    lambda item: (item[key] == column_name),
                    items,
                ),
                None,
            ):
                cell.value = item["value"]
            else:
                cell.value = "-"

    @staticmethod
    def get_category_data_by_dow(
        count: modelCount,
        section=None,
        categoryid=None,
        lane=None,
        direction=None,
        start=None,
        end=None,
    ) -> "ValuesQuerySet[models.CountDetail, Any]":
        if not start:
            start = count.start_process_date
        if not end:
            end = count.end_process_date + timedelta(days=1)
        start, end = tuple([utils.to_time_aware_utc(d) for d in (start, end)])

        qs = CountDetail.objects.filter(
            id_lane__id_section=section,
            id_category=categoryid,
            timestamp__gte=start,
            timestamp__lt=end,
        )

        if count is not None:
            qs = qs.filter(id_count=count)

        if lane is not None:
            qs = qs.filter(id_lane=lane)

        if direction is not None:
            qs = qs.filter(id_lane__direction=direction)

        qs = (
            qs.annotate(week_day=ExtractIsoWeekDay("timestamp"))
            .values("week_day", "times")
            .annotate(value=Sum("times"))
            .values("week_day", "value")
            .values_list("week_day", "value")
        )
        print(
            "yearly_report_bike.py : get_category_data_by_dow - qs.query=",
            str(qs.query),
        )

        return qs

    def run(self):
        print(f"{datetime.now()}: YRB_run - begin... ({self.path_to_output_dir})")
        current_dir = path.dirname(path.abspath(__file__))
        template = path.join(current_dir, "template_yearly_bike.xlsx")
        workbook = load_workbook(filename=template)

        """ Data_count """

        ws = workbook["Data_count"]

        section = Section.objects.get(id=self.section_id)

        def render_section_dist(value: Union[str, Decimal, None]) -> str:
            if value is None or value == "NA":
                return "NA"
            if isinstance(value, str):
                return str(round(int(value)))
            if isinstance(value, Decimal):
                return str(round(value))
            raise ValueError(value)

        section_start_dist = render_section_dist(section.start_dist)
        section_end_dist = render_section_dist(section.end_dist)

        ws[
            "B3"
        ] = f"""
            Poste de comptage : {section.id}
            Axe : {section.owner}:{section.road}{section.way}
            PR {section.start_pr} + {section_start_dist} m à PR {section.end_pr} + {section_end_dist} m
        """

        ws["B4"] = "Periode de comptage du 01/01/{0} au 31/12/{0}".format(self.year)
        ws["B5"] = "Comptage {}".format(self.year)

        # Get one count for the section and the year to get the base data
        count_detail = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        if not count_detail.exists():
            QgsMessageLog.logMessage(
                f"{datetime.now()} - Aucun conmptage pour cette année ({self.year}) et cette section ({self.section_id})",
                "Comptages",
                Qgis.Info,
            )
            return

        count = count_detail[0].id_count

        ws["B6"] = "Type de capteur : {}".format(count.id_sensor_type.name)
        ws["B7"] = "Modèle : {}".format(count.id_model.name)
        ws["B8"] = "Classification : {}".format(count.id_class.name)
        ws["B9"] = "Comptage véhicule par véhicule"
        ws["B12"] = "Remarque : {}".format(count.remarks)

        lanes = Lane.objects.filter(id_installation=count.id_installation)

        ws["B13"] = lanes[0].direction_desc
        if len(lanes) > 1:
            ws["B14"] = lanes[1].direction_desc

        ws["B11"] = lanes[0].id_section.place_name

        """ Data_year """

        ws = workbook["Data_year"]
        row_offset = 4
        column_offset = 1

        data = self.total_runs_by_day()
        row = row_offset
        for i in data:
            ws.cell(row=row, column=column_offset, value=i["date"])
            ws.cell(row=row, column=column_offset + 1, value=i["daily_runs"])
            row += 1

        """ Data_week """

        ws = workbook["Data_week"]
        row_offset = 4
        column_offset = 2

        data = self.tjms_total_runs_by_day_of_week().values()
        row = row_offset
        for i in data:
            ws.cell(row=row, column=column_offset, value=i["runs"])
            ws.cell(row=row, column=column_offset + 3, value=i["tjm"])
            row += 1

        """ Data_hour """

        ws = workbook["Data_hour"]

        # Data hour > Whole weeks
        print_area = ws["C5:D28"]
        data = self.total_runs_by_hour_and_direction(directions=(1, 2))
        for hour, row in enumerate(print_area, 0):
            for direction, cell in enumerate(row, 1):
                if hour not in data or direction not in data[hour]:
                    cell.value = 0
                else:
                    cell.value = data[hour][direction]["runs"]

        # Data hour > Weekends only
        print_area = ws["C37:D60"]
        data = self.total_runs_by_hour_and_direction(directions=(1, 2), weekdays=(6, 7))
        for hour, row in enumerate(print_area, 0):
            for direction, cell in enumerate(row, 1):
                if hour not in data or direction not in data[hour]:
                    cell.value = 0
                else:
                    cell.value = data[hour][direction]["runs"]

        # Data hour > Only dir 1
        print_area = ws["B69:H92"]
        data = self.total_runs_by_hour_one_direction(1)
        for hour, row in enumerate(print_area, 0):
            for day, cell in enumerate(row, 1):
                if day not in data or hour not in data[day]:
                    cell.value = 0
                else:
                    cell.value = data[day][hour]

        # Data hour > Only dir 2
        print_area = ws["B101:H124"]
        data = self.total_runs_by_hour_one_direction(2)
        for hour, row in enumerate(print_area, 0):
            for day, cell in enumerate(row, 1):
                if day not in data or hour not in data[day]:
                    cell.value = 0
                else:
                    cell.value = data[day][hour]

        """ Data_yearly_stats """

        ws = workbook["Data_yearly_stats"]
        print_area = ws["B2:G8"]
        data = self.count_details_by_various_criteria(self, count)
        column_names = (
            "VELO",
            "MONO",
            "SHORT",
            "SPECIAL",
            "MULTI",
            "day_or_month_or_weekend",
        )
        row_names = (
            "total_runs_in_year",
            "busiest_date_row",
            "least_busy_date_row",
            "busiest_month_row",
            "least_busy_month_row",
            "total_runs_busiest_hour_weekday",
            "total_runs_busiest_hour_weekend",
        )
        for row_idx, row_name in enumerate(row_names, 0):
            row = print_area[row_idx]
            YearlyReportBike.write_to_row(
                row_name=row_name,
                row=row,
                data=data,
                key="category_name",
                column_names=column_names,
            )

        """ Section Passages saisonniers pour chaque catégories Adrien"""
        row_offset = 22
        column_offset = 2

        data = self.count_details_by_season(self, count)
        for saison in data:
            for i, season in enumerate(self.seasons):
                if saison == season:
                    for j, cat in enumerate(column_names):
                        if cat in data[saison]:
                            ws.cell(
                                column=j + column_offset,
                                row=i + row_offset,
                                value=data[saison][cat],
                            )

        """ Section Passages mensuels Adrien"""
        row_offset = 40
        column_offset = 1

        data = self.runs_by_weekday_and_month()
        for i in data:
            ws.cell(
                column=i["week_day"] + column_offset,
                row=i["month"] + row_offset,
                value=i["daily_runs"],
            )

        row_offset = 60
        data = self.nb_weekday_by_month()
        for mois in data:
            for jours in data[mois]:
                ws.cell(
                    column=jours + column_offset,
                    row=mois + row_offset,
                    value=data[mois][jours],
                )

        """ Data_category """
        """ Direction Mario"""
        ws = workbook["Data_category"]

        start = datetime(self.year, 1, 1)
        end = datetime(self.year + 1, 1, 1)
        section = Section(self.section_id)

        categories = (
            Category.objects.filter(countdetail__id_count=count)
            .distinct()
            .order_by("code")
        )
        print(
            f"yearly_report_bike.py : Data_category, Direction Mario, categories.query={str(categories.query)}"
        )

        # Direction 1
        row_offset = 5
        col_offset = 1
        for category in categories:
            res = self.get_category_data_by_dow(
                count=None,
                section=self.section_id,
                categoryid=category.id,
                lane=None,
                direction=1,
                start=start,
                end=end,
            )

            for row in res:
                row_num = row_offset + category.code
                col_num = col_offset + row[0]
                val = ws.cell(row_num, col_num).value
                value = (
                    val + row[1] if isinstance(val, (int, float)) else row[1]
                )  # Add to previous value because with class convertions multiple categories can converge into a single one

                ws.cell(row=row_num, column=col_num, value=value)

        # Direction 2
        if len(section.lane_set.all()) == 2:
            row_offset = 29
            col_offset = 1
            for category in categories:
                res = self.get_category_data_by_dow(
                    None,
                    section,
                    categoryid=category.id,
                    direction=2,
                    start=start,
                    end=end,
                )

                for row in res:
                    row_num = row_offset + category.code
                    col_num = col_offset + row[0]
                    val = ws.cell(row_num, col_num).value
                    value = (
                        val + row[1] if isinstance(val, (int, float)) else row[1]
                    )  # Add to previous value because with class convertions multiple categories can converge into a single one

                    ws.cell(row=row_num, column=col_num, value=value)

        # Section
        row_offset = 53
        col_offset = 1
        for category in categories:
            res = self.get_category_data_by_dow(
                count=None,
                section=self.section_id,
                categoryid=category.id,
                lane=None,
                direction=None,
                start=start,
                end=end,
            )

            for row in res:
                row_num = row_offset + category.code
                col_num = col_offset + row[0]
                val = ws.cell(row_num, col_num).value
                value = (
                    val + row[1] if isinstance(val, (int, float)) else row[1]
                )  # Add to previous value because with class convertions multiple categories can converge into a single one

                ws.cell(row=row_num, column=col_num, value=value)

        # Save the file
        output = path.join(
            self.path_to_output_dir, "{}_{}_r.xlsx".format(self.section_id, self.year)
        )

        workbook.save(filename=output)
        print(f"{datetime.now()}: YRB_run - end: Saved report to {output}")
