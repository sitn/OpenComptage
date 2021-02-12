import os
import datetime
from openpyxl import load_workbook

from qgis.PyQt.QtCore import QDate
from comptages.core.settings import Settings

from comptages.data.data_loader import DataLoader
from comptages.data.count_data import CountData
from comptages.core.layers import Layers


class YearlyReportCreator():
    def __init__(self, file_path, layers, year, section_id):
        self.file_path = file_path
        self.layers = layers
        self.year = year
        self.section_id = section_id

    def run(self):

        # TODO: get the list of all the count_is of the section for the year

        counts = self.layers.get_counts_of_section_by_year(
            self.section_id, self.year)

        merged_count_data = CountData()

        for count in counts:
            count_id = count.attribute('id')
            data_loader = DataLoader(
                count_id, self.section_id,
                Layers.IMPORT_STATUS_DEFINITIVE)
            merged_count_data = data_loader.load(merged_count_data)

        self._export_report(merged_count_data)

    def _export_report(self, count_data):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        template = os.path.join(current_dir, 'template_yearly.xlsx')
        wb = load_workbook(filename=template)

        self._set_data_count(wb, count_data)
        self._set_data_day(wb, count_data)
        self._set_data_speed(wb, count_data)

        self._set_data_category(wb, count_data)
        self._remove_useless_sheets(wb, count_data)

        # Save the file
        output = os.path.join(
            self.file_path, '{}_{}_r.xlsx'.format(self.section_id, self.year))

        wb.save(filename=output)

    def _set_data_count(self, workbook, count_data):

        ws = workbook['Data_count']

        ws['B3'] = ('Poste de comptage : {}  Axe : {}:{}{}  '
                    'PR {} + {} m à PR {} + {} m').format(
                        self.section_id,
                        count_data.attributes['owner'],
                        count_data.attributes['road'],
                        count_data.attributes['way'],
                        count_data.attributes['start_pr'],
                        int(round(count_data.attributes['start_dist'])),
                        count_data.attributes['end_pr'],
                        int(round(count_data.attributes['end_dist'])))

        ws['B4'] = 'Periode de comptage du 01/01/{0} au 31/12/{0}'.format(
            self.year)

        ws['B5'] = 'Comptage {}'.format(self.year)

        ws['B6'] = 'Type de capteur : {}'.format(
            count_data.attributes['sensor_type'])
        ws['B7'] = 'Modèle : {}'.format(
            count_data.attributes['model'])
        ws['B8'] = 'Classification : {}'.format(
            count_data.attributes['class'])

        ws['B9'] = 'Comptage véhicule par véhicule'
        if count_data.attributes['aggregate']:
            ws['B9'] = 'Comptage par interval'

        ws['B11'] = count_data.attributes['place_name']

        ws['B12'] = 'Remarque : {}'.format(
            count_data.attributes['remarks']
            if type(count_data.attributes['remarks']) == str else '')

        ws['B13'] = count_data.attributes['dir1']

        if 'dir2' in count_data.attributes:
            ws['B14'] = count_data.attributes['dir2']

    def _set_data_day(self, workbook, count_data):
        ws = workbook['Data_day']

        # TODO: aggregate days
        days = count_data.day_data[0:7]

        day_cols_tot = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
        section_start_cell = 5
        coefficient_cell = 31
        dir1_start_cell = 35
        dir1_light_cell = 61
        dir1_heavy_cell = 62
        dir2_start_cell = 66
        dir2_light_cell = 92
        dir2_heavy_cell = 93
        for i, day_data in enumerate(days):
            for j in range(24):
                ws['{}{}'.format(day_cols_tot[i], j+section_start_cell)] = \
                    day_data.hour_data[j].total()
                ws['{}{}'.format(day_cols_tot[i], j+dir1_start_cell)] = \
                    day_data.hour_data[j].total(0)
                ws['{}{}'.format(day_cols_tot[i], j+dir2_start_cell)] = \
                    day_data.hour_data[j].total(1)

            ws['{}{}'.format(day_cols_tot[i], coefficient_cell)] = \
                '{}%'.format(day_data.monthly_coefficient)
            ws['{}{}'.format(day_cols_tot[i], dir1_light_cell)] = \
                day_data.light_vehicles(0)
            ws['{}{}'.format(day_cols_tot[i], dir2_light_cell)] = \
                day_data.light_vehicles(1)
            ws['{}{}'.format(day_cols_tot[i], dir1_heavy_cell)] = \
                day_data.heavy_vehicles(0)
            ws['{}{}'.format(day_cols_tot[i], dir2_heavy_cell)] = \
                day_data.heavy_vehicles(1)

    def _set_data_speed(self, workbook, count_data):
        ws = workbook['Data_speed']

        speed_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H',
                      'I', 'J', 'K', 'L', 'M', 'N']
        dir1_start_cell = 5
        dir2_start_cell = 33
        average_class_speeds_row = 60

        average_class_speeds = []
        if count_data.attributes['aggregate']:
            average_class_speeds = [12.5, 22.5, 35, 45,
                                    55, 65, 75, 85, 95,
                                    105, 115, 125]

            for i, speed in enumerate(average_class_speeds):
                ws['{}{}'.format(
                    speed_cols[i], average_class_speeds_row)] = speed

        #days_idx = [x for x in range(start_day, end_day+1)]
        days_idx = [x for x in range(8)]
        dir1 = count_data.speed_cumulus(0, days=days_idx)
        dir2 = count_data.speed_cumulus(1, days=days_idx)

        start_timestamp = datetime.date(self.year, 1, 1).strftime('%Y-%m-%d')

        end_timestamp = (datetime.date(self.year, 1, 1).strftime('%Y-%m-%d'))

        for i, hour in enumerate(dir1):
            for j, speed in enumerate(hour):
                ws['{}{}'.format(speed_cols[j], i+dir1_start_cell)] = speed

            # if not count_data.attributes['aggregate']:
            #     # Average and characteristic speed
            #     char_speed = self.layers.get_characteristic_speeds(
            #         self.count_id, i, 1, start_timestamp, end_timestamp,
            #         self.section_id)
            #     ws['P{}'.format(i+dir1_start_cell)] = char_speed[0]
            #     ws['Q{}'.format(i+dir1_start_cell)] = char_speed[1]
            #     ws['R{}'.format(i+dir1_start_cell)] = char_speed[2]
            #     ws['S{}'.format(i+dir1_start_cell)] = char_speed[3]

        for i, hour in enumerate(dir2):
            for j, speed in enumerate(hour):
                ws['{}{}'.format(speed_cols[j], i+dir2_start_cell)] = speed

            # if not count_data.attributes['aggregate']:
            #     # Average and characteristic speed
            #     char_speed = self.layers.get_characteristic_speeds(
            #         self.count_id, i, 2, start_timestamp, end_timestamp,
            #         self.section_id)
            #     ws['P{}'.format(i+dir2_start_cell)] = char_speed[0]
            #     ws['Q{}'.format(i+dir2_start_cell)] = char_speed[1]
            #     ws['R{}'.format(i+dir2_start_cell)] = char_speed[2]
            #     ws['S{}'.format(i+dir2_start_cell)] = char_speed[3]

    def _set_data_category(self, workbook, count_data):
        if self._translate_class_name(count_data.attributes['class']) == 'Volume':
            return

        ws = workbook['Data_category']

        cat_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H',
                    'I', 'J', 'K']

        dir1_start_cell = 5
        dir2_start_cell = 33

        days_idx = [x for x in range(8)]
        dir1 = count_data.category_cumulus(0, days=days_idx)
        dir2 = count_data.category_cumulus(1, days=days_idx)

        for i, hour in enumerate(dir1):
            # The report is configurated for SWISS7 or SWISS10 data.
            # If the data is in another class, we need to translate it to
            # SWISS7 or SWISS10
            hour = self._translate_class_hour(count_data, hour)
            for j, cat in enumerate(hour):
                # Don't consider cat 0 (TRASH) in the report
                if j == 0:
                    continue

                ws['{}{}'.format(cat_cols[j-1], i+dir1_start_cell)] = cat

        for i, hour in enumerate(dir2):
            hour = self._translate_class_hour(count_data, hour)
            for j, cat in enumerate(hour):
                # Don't consider cat 0 (TRASH) in the report
                if j == 0:
                    continue

                ws['{}{}'.format(cat_cols[j-1], i+dir2_start_cell)] = cat

    def _remove_useless_sheets(self, workbook, count_data):

        class_name = self._translate_class_name(count_data.attributes['class'])

        if class_name == 'SWISS10':
            workbook.remove_sheet(workbook['SWISS7_H'])
            workbook.remove_sheet(workbook['SWISS7_G'])
        elif class_name == 'SWISS7':
            workbook.remove_sheet(workbook['SWISS10_H'])
            workbook.remove_sheet(workbook['SWISS10_G'])
        elif class_name == 'Volume':
            workbook.remove_sheet(workbook['SWISS7_H'])
            workbook.remove_sheet(workbook['SWISS7_G'])
            workbook.remove_sheet(workbook['SWISS10_H'])
            workbook.remove_sheet(workbook['SWISS10_G'])

        if count_data.attributes['aggregate']:
            workbook.remove_sheet(workbook['Vit_Hd'])
        else:
            workbook.remove_sheet(workbook['Vit_H'])

    def _translate_class_hour(self, count_data, hour):
        """Convert a class to another one e.g. FHWA13
        should be converted in SWISS7 in order to fill the
        report cells"""

        if count_data.attributes['class'] == 'SWISS10':
            return hour

        if count_data.attributes['class'] == 'SWISS7':
            return hour

        if count_data.attributes['class'] == 'ARX Cycle':
            new_hour = [0] * 7
            return new_hour

        if count_data.attributes['class'] == 'FHWA13':
            new_hour = [0] * 8

            new_hour[0] = hour[0]
            new_hour[1] = hour[4]
            new_hour[2] = hour[1]
            new_hour[3] = hour[2]
            new_hour[4] = hour[3]
            new_hour[5] = hour[5] + hour[6] + hour[7]
            new_hour[6] = hour[11] + hour[12]
            new_hour[7] = hour[8] + hour[9] + hour[10] + hour[13] + hour[14]
            return new_hour

    def _translate_class_name(self, name):

        if name == 'SWISS10':
            return name

        if name == 'SWISS7':
            return name

        if name == 'ARX Cycle':
            return name

        if name == 'FHWA13':
            return 'SWISS7'

        if name is None:
            return 'Volume'